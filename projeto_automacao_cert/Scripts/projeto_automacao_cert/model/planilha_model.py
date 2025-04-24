# model/planilha_model.py

from openpyxl import load_workbook
from typing import List


class PlanilhaModel:
    def __init__(self, caminho_arquivo: str):
        self.caminho = caminho_arquivo
        self.wb = load_workbook(caminho_arquivo)
        self.ws = self.wb.active

    def obter_cnpjs(self, linha_excel) -> List[str]:
        """
        Lê os CNPJs da coluna B (linha 2 em diante).
        """
        cnpjs = []
        for linha in self.ws.iter_rows(min_row=linha_excel, max_col=2, values_only=True):
            cnpj = linha[1]
            if cnpj:
                cnpjs.append(str(cnpj).strip())
        return cnpjs
    
    def obter_razao_social(self, linha_excel) -> List[str]:
        """
        Lê os nomes das empresas da coluna A (linha `linha_excel` em diante).
        """
        lista_razoes = []
        for linha in self.ws.iter_rows(min_row=linha_excel, max_col=1, values_only=True):
            nome = linha[0]
            if nome:
                lista_razoes.append(str(nome).strip())
        return lista_razoes

    def atualizar_downloads_GFD(self, linha: int, status_gfd: str):
        """
        Atualiza a colunas D (GFD).
        """
        self.ws.cell(row=linha, column=4, value=status_gfd)     # Coluna D
        self.wb.save(self.caminho)
    
    def atualizar_downloads_detalhe_da_guia(self, linha: int, status_detalhe: str):
        """
        Atualiza a coluna E (Detalhe da Guia).
        """
        self.ws.cell(row=linha, column=5, value=status_detalhe) # Coluna E
        self.wb.save(self.caminho)
    
    

    def atualizar_situacao(self, linha: int, mensagem: str):
        """
        Atualiza a coluna F com a situação final (mensagem de sucesso ou erro).
        """
        self.ws.cell(row=linha, column=6, value=mensagem)       # Coluna F
        self.wb.save(self.caminho)
    '''
    
    '''