from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from time import sleep
import os


# Função para sair e fechar o navegador
def sair_e_fechar(
    page,
    browser,
):
    try:
        # Tira um screenshot da ultima pagina aberta
        page.screenshot(
            path=os.path.join(
                download_directory_screenshot_ultima_pag_open, "screenshot.png"
            ),
            full_page=True,
        )
        # Clique no botão de sair
        page.click("svg.feather.feather-log-out")
        sleep(2)
        print("Logout realizado com sucesso!")
    except Exception as e:
        print(f"Erro ao tentar sair: {e}")
    finally:
        # Fecha o navegador
        browser.close()


def aguarda(tempo=1):
    sleep(tempo)


def carregar_p(page):
    # Aguarda a navegação
    page.wait_for_load_state("load")

print('Informe a data da competencia que que sera feita a busca da NFSe \nEX: 03-2025')
data_compentencia = input('Competencia: ')
# Código principal
with sync_playwright() as p:
    browser = p.chromium.launch(
        headless=True
    )  # Altere para False para visualizar o navegador
    context = browser.new_context(
        accept_downloads=True, base_url="https://iss.itatiba.sp.gov.br/"
    )
    page = context.new_page()

    # Defina o seletor
    seletor_do_link_exporta_nfse = (
        'a.nav-sub-link[title="Efetue a exportação de suas notas para xml.."]'
    )
    # Configurações iniciais
    expected_url = "https://iss.itatiba.sp.gov.br/NovoContador"
    expected_url_domicilio_tributario_eletronico = "https://iss.itatiba.sp.gov.br/Dte"
    expected_url_exportacao_nfse = "https://iss.itatiba.sp.gov.br/ExportarNfse"

    # Caminho para salvar downloads da ultima pagina aberta
    download_directory_screenshot_ultima_pag_open = (
        r"C:\Users\Jota11\AppData\Local\anaconda3\envs\automacao_relatorio_nfse\src\img\ultima-pag-open"
    )
    os.makedirs(download_directory_screenshot_ultima_pag_open, exist_ok=True)

    # Caminho para salvar downloads das notas
    download_directory_nfse = r"C:\Users\Jota11\AppData\Local\anaconda3\envs\rpa_baixa_nfse_e_transcreve\src\download"
    os.makedirs(download_directory_nfse, exist_ok=True)

    # Caminho para salvar print da tela que lista notas
    download_directory_screenshot_tela_notas = (
        r"C:\Users\Jota11\AppData\Local\anaconda3\envs\automacao_relatorio_nfse\src\img\emp-tela-notas"
    )
    os.makedirs(download_directory_screenshot_tela_notas, exist_ok=True)

    # Caminho para salvar print da tela da verificacao de empresa valida
    donwload_directory_screenshot_empresa_habilitada = (
        r"C:\Users\Jota11\AppData\Local\anaconda3\envs\automacao_relatorio_nfse\src\img\emp-habilitada"
    )
    os.makedirs(donwload_directory_screenshot_empresa_habilitada, exist_ok=True)

    # Caminho do arquivo Excel que contem as empresas
    caminho_arquivo = r"C:\Users\Jota11\AppData\Local\anaconda3\envs\rpa_baixa_nfse_e_transcreve\Planilhas\Empresas.xlsx"
    # Carregar o arquivo Excel
    workbook = load_workbook(caminho_arquivo)
    # Selecionar a primeira planilha
    sheet = workbook.active

    # Definir a coluna a ser analisada (ex.: Coluna A é 1, Coluna B é 2)
    coluna = 2  # Número da coluna desejada
    contador = 0
    for linha in range(1, sheet.max_row + 1):  # De 1 até a última linha
        valor = sheet.cell(row=linha, column=coluna).value
        if (
            valor is not None and str(valor).strip() != ""
        ):  # Verifica se a célula não está vazia
            contador += 1

    print(f"A coluna {coluna} contém {contador} linhas com valores.")

    contador += 1
    linha_coletora_excel = 2

    while linha_coletora_excel != contador:
        try:

            # Navegar para a página de login
            page.goto("Account/LogonContador")
            # Realizar o login
            page.locator("button").nth(0).click()
            page.fill("input#Cnpj", "05.369.682/0001-28")
            page.fill("input#Email", "fiscal2@jotacontabil.com.br")
            page.fill("input#Senha", "00229527")
            page.click("button.btn.btn-brand-02")
            aguarda()
            # Aguarda a navegação
            carregar_p(page)

            # Verifica se está na URL esperada
            if page.url == expected_url:
                print("Login bem-sucedido! Estamos na página esperada.")
                # Obter o valor da célula na primeira linha e primeira coluna (A1)
                obter_id = sheet.cell(row=linha_coletora_excel, column=2).value
                # Obter o valor da célula na primeira linha e primeira coluna (A1)
                obter_cnpj = sheet.cell(row=linha_coletora_excel, column=4).value

                # Inserindo um CNPJ para iniciar a pesquisa
                page.fill("input#Filtro_CnpjCpf", obter_cnpj)
                # Clicar no filtro situacao para filtrar apenas empresas ATIVAS
                page.click("#Filtro_Situacao")
                page.select_option("#Filtro_Situacao", value="1")

                # Clicando em Consultar
                page.click('input[value="Consultar"]')

                # Aguarda a navegação
                aguarda()
                carregar_p(page)
                # Tirar Print
                page.screenshot(
                    path=os.path.join(
                        donwload_directory_screenshot_empresa_habilitada,
                        f"screenshot_id_{obter_id}.png",
                    ),
                    full_page=True,
                )

            else:
                print("A URL é diferente do esperado. Tentando novamente...")
                continue

            # Localiza todos os radio buttons dentro das células da tabela
            radio_buttons = page.locator('td.text-center input[type="radio"]')

            # Flag para verificar se encontramos um botão de rádio visível
            encontrou_radio_visivel = False

            # Localiza todas as linhas da tabela
            linhas_tbody = page.locator("tbody > tr")

            # Itera sobre as linhas da tabela
            for i in range(linhas_tbody.count()):
                linha = linhas_tbody.nth(i)

                # Verifica se a linha contém o texto "Ativa"
                if linha.locator("td").filter(has_text="Ativa").count() > 0:
                    # Se a linha contém "Ativa", clica no botão de rádio
                    radio_button = linha.locator("input[type='radio']")
                    if radio_button.is_visible():
                        radio_button.click()
                        print(f"Botão de rádio na linha {i} (Ativa) clicado.")
                        carregar_p(page)  # Realiza ação pós-clique
                        # Atualiza o Excel para indicar que a empresa está inativa
                        sheet.cell(row=linha_coletora_excel, column=6).value = "A"
                        workbook.save(caminho_arquivo)
                        encontrou_radio_visivel = True

            # Se não encontrou nenhum botão de rádio visível
            if not encontrou_radio_visivel:
                # Atualiza o Excel para indicar que a empresa está inativa
                sheet.cell(row=linha_coletora_excel, column=6).value = "I"
                workbook.save(caminho_arquivo)
                linha_coletora_excel += 1  # Avança para a próxima linha no Excel
                continue
            
            # Verifica se chegou à página Domicílio Tributário Eletrônico
            if page.url == expected_url_domicilio_tributario_eletronico:
                print("Estamos na página de Domicílio Tributário Eletrônico.")
                # Clique no menu "Operações Fiscais"
                page.click("#navbarMenu > ul > li:nth-child(4) > a")
                print("click enviado")
                print("Aguardando seletor....")
                aguarda()
                carregar_p(page)

                # Aguarde o menu Operacoes Fiscais carregar, se necessário
                # page.wait_for_selector('a.nav-sub-link[title="Efetue a exportação de suas notas para xml.."]')

                # Verifica se o seletor está visível
                if page.is_visible(seletor_do_link_exporta_nfse):
                    # Se o seletor estiver visível, clica no elemento
                    page.click(seletor_do_link_exporta_nfse)
                    print("Elemento clicado.")
                else:
                    # Caso contrário, continua a execução
                    print("Elemento não está visível, seguindo em frente.")
                    # Edita a culuna 3 do excel adicionado a string "Download-efetuado"
                    sheet.cell(row=linha_coletora_excel, column=5).value = (
                        "ExportaNfseNaoHabilitado"
                    )
                    workbook.save(caminho_arquivo)
                    print("Excel editado e salvo")
                    linha_coletora_excel += 1
                    continue

            else:
                print("A URL é diferente do esperado. Tentando novamente...")
                continue

            # Verifica se chegou à página Exportação NFSE
            if page.url == expected_url_exportacao_nfse:
                print("Estamos na página de Exportação NFSE.")
                # Modificando a data para filtrar a NFSE
                page.click("#Filtro_tipoDataAvancada")
                aguarda()
                page.select_option("#Filtro_tipoDataAvancada", value="1")
                aguarda()
                
                page.fill("#Filtro_dataInicialEmissao", data_compentencia)
                aguarda()
                page.fill("#Filtro_dataFinalEmissao", data_compentencia)
                aguarda()
                page.click("#btnConsultar")
                aguarda()
                carregar_p(page)

                # Tirar Print da tela de notas
                page.screenshot(
                    path=os.path.join(
                        download_directory_screenshot_tela_notas,
                        f"screenshot_id_{obter_id}.png",
                    ),
                    full_page=True,
                )

                # Este seletor corresponte a 1 Nota para ser baixada entao se contem este seletor contem notas a ser baixadas
                seletorBotaoNfse = "#example > tbody > tr:nth-child(1) > td.dt-body-center > input[type=checkbox]"

                if page.is_visible(seletorBotaoNfse):
                    # Sequência do código se o seletor existir
                    print(
                        "O seletor existe entao contem notas, prosseguindo com o código."
                    )
                    page.click("#example-select-all")
                    carregar_p(page)
                    page.click("#versaoSelecionada")
                    carregar_p(page)
                    page.select_option("#versaoSelecionada", value="2")
                    carregar_p(page)
                    aguarda()
                    with page.expect_download() as download_info:
                        page.click("#bntGerar")  # Clique no botão que gera o download
                        carregar_p(page)
                        sleep(3)
                    download = download_info.value
                    # Salva o arquivo no diretório especificado
                    download.save_as(
                        os.path.join(
                            download_directory_nfse,
                            f"{obter_id}-",
                            download.suggested_filename,
                        )
                    )
                    print(
                        f"Download salvo em: {os.path.join(download_directory_nfse, download.suggested_filename)}"
                    )

                    # Edita a culuna 3 do excel adicionado a string "Download-efetuado"
                    sheet.cell(row=linha_coletora_excel, column=5).value = (
                        "Download-efetuado"
                    )
                    workbook.save(caminho_arquivo)
                    print("Excel editado e salvo")

                    linha_coletora_excel += 1
                    page.goto("/NovoContador")
                    carregar_p(page)
                    aguarda()

                else:
                    # Caso o seletor não exista
                    print("Não contém notas para serem baixadas.")
                    # Edita a culuna 3 do excel adicionado a string "Download-efetuado"
                    sheet.cell(row=linha_coletora_excel, column=5).value = "Sem notas"
                    workbook.save(caminho_arquivo)
                    print("Excel editado e salvo")
                    linha_coletora_excel += 1
                # Enviar um Go To Back ate encontrar a pagina de consulta de empresas
                carregar_p(page)
                aguarda()
                continue

            else:
                print("A URL é diferente do esperado. Tentando novamente...")
                continue

        except Exception as e:
            print(f"Ocorreu um erro: {e}")
            continue

        break