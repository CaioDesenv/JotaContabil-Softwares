# processamento_planilha.py
import openpyxl
import smtplib
import traceback
from email.message import EmailMessage
from smtp_config import SMTP_CONFIGURACAO  # Importa a configuração centralizada do SMTP
from email_formatter import EmailFormatter  # Importa o formatador de emails

class ProcessamentoPlanilha:
    """
    Classe responsável por processar a planilha Excel e enviar notificações por e-mail
    caso sejam identificadas ocorrências fiscais (quando o status da coluna 5 não for 'Ativa').
    """
    def __init__(self, caminho_planilha: str, smtp_config: dict = SMTP_CONFIGURACAO):
        """
        Inicializa a classe com o caminho do arquivo Excel, as configurações do servidor SMTP
        e carrega o workbook e a planilha ativa.
        """
        self.caminho_planilha = caminho_planilha
        self.smtp_config = smtp_config
        self.destinatario = "info@jotacontabil.com.br, rh1@jotacontabil.com.br, rafael@jotacontabil.com.br"
        
        # Carrega o workbook e a planilha ativa no momento da instância
        try:
            self.workbook = openpyxl.load_workbook(self.caminho_planilha)
            self.sheet = self.workbook.active
        except Exception as erro:
            print(f"Erro ao carregar a planilha: {erro}")
            self.workbook = None
            self.sheet = None

    def atualizar_resultado(self, linha: int, resultado: str, ocorrencia: str) -> None:
        """
        Atualiza as colunas de resultado e ocorrência na planilha para a linha indicada.
        """
        if self.sheet:
            self.sheet.cell(row=linha, column=4, value=resultado)
            self.sheet.cell(row=linha, column=5, value=ocorrencia)
            self.workbook.save(self.caminho_planilha)
            print(f"Planilha atualizada para a linha {linha}.")
        else:
            print("Planilha não carregada.")

    def processar_planilha(self) -> dict:
        """
        Percorre todas as linhas da planilha e retorna um dicionário com as linhas 
        onde o status na coluna 5 não é 'Ativa'.
        """
        ocorrencias = {}
        if not self.sheet:
            print("Nenhuma planilha carregada para processamento.")
            return ocorrencias

        try:
            for numero_linha, linha in enumerate(self.sheet.iter_rows(min_row=1, values_only=True), start=1):
                # A coluna 5 é o índice 4 (lembrando que a contagem começa em 0)
                valor_status = linha[4]
                if valor_status is None or str(valor_status).strip() != "Ativa":
                    ocorrencias[numero_linha] = linha
            return ocorrencias
        except Exception as erro:
            print(f"Erro ao processar a planilha: {erro}")
            return ocorrencias

    def enviar_email(self, dados: dict) -> None:
        """
        Caso existam ocorrências, monta e envia um e-mail para o destinatário informado,
        listando as empresas com ocorrências identificadas.
        """
        if not dados:
            print("Nenhuma ocorrência encontrada. E-mail não enviado.")
            return

        # Usar o formatador de email para criar o conteúdo
        email_formatado = EmailFormatter.formatar_email_ocorrencias_cadesp(dados)
        
        # Configurar a mensagem de email
        mensagem = EmailMessage()
        mensagem['Subject'] = email_formatado["assunto"]
        mensagem['From'] = self.smtp_config.get('remetente')
        mensagem['To'] = self.destinatario
        
        # Adicionar as duas versões do conteúdo (HTML e texto)
        mensagem.set_content(email_formatado["texto"])
        mensagem.add_alternative(email_formatado["html"], subtype='html')
        
        # Enviar o email
        try:
            with smtplib.SMTP_SSL(self.smtp_config.get('host'), self.smtp_config.get('port')) as servidor:
                servidor.login(self.smtp_config.get('usuario'), self.smtp_config.get('senha'))
                servidor.send_message(mensagem)
                print(f"E-mail enviado com sucesso para {self.destinatario}")
        except Exception as erro:
            print(f"Erro ao enviar e-mail via SMTP_SSL: {erro}")
            # Registrar detalhes adicionais do erro para facilitar o diagnóstico
            print(traceback.format_exc())

if __name__ == "__main__":
    caminho_excel = r"C:\Users\Jota11\AppData\Local\anaconda3\envs\automacao_consulta_icms_cadesp\src\planilha\Relatorio-de-consulta-CADESP.xlsx"
    processador = ProcessamentoPlanilha(caminho_excel)
    ocorrencias = processador.processar_planilha()
    if ocorrencias:
        processador.enviar_email(ocorrencias)
    else:
        print("Nenhuma ocorrência identificada.")
